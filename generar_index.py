# -*- coding: utf-8 -*-
"""
Genera HTML (index/live) incrustando DATA.rows desde el Excel,
y las SUSTITUCIONES SIEMPRE desde un CSV fijo (ignora hoja "Sustituciones").
"""

import sys, os, json
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any
from datetime import date, datetime as _dt

try:
    import openpyxl
except Exception as e:
    print("❌ Falta openpyxl. Instala dependencias con: pip install -r requirements.txt")
    sys.exit(1)

# === CONFIG ===
# Excel (cuadrantes) - intenta varias rutas conocidas
EXCEL_CANDIDATOS = [
    r"C:\Users\comun\OneDrive\02. Comp. Min Recepción\3. Turnos\Plantilla Cuadrante con Sustituciones v.6.0.xlsx",
    r"C:\Users\comun\Documents\Turnos web\Plantilla Cuadrante con Sustituciones v.6.0.xlsx",
]
# FORZAR lectura de sustituciones desde este CSV (ignora Excel siempre)
CSV_SUSTITUCIONES = r"C:\Users\comun\Documents\Turnos web\sustituciones_diagnostico.csv"

# Nombre del archivo HTML objetivo se define en cada script (INDEX_HTML / LIVE_HTML).

def localizar_excel() -> str:
    for p in EXCEL_CANDIDATOS:
        if os.path.exists(p):
            return p
    print("❌ No encuentro el Excel. Revisa EXCEL_CANDIDATOS en el script.")
    sys.exit(1)

def to_safe_str(v):
    if v is None:
        return ""
    if isinstance(v, (_dt, date)):
        try:
            return v.strftime("%d/%m/%Y")
        except Exception:
            return str(v)
    if isinstance(v, float) and str(v) == "nan":
        return ""
    return str(v).strip()

def leer_hoteles_y_rows(desde_excel: str) -> List[Dict[str, Any]]:
    """
    Lee todas las hojas de hoteles y devuelve filas con:
    { 'hotel','semana','empleado','lunes','martes','miercoles','jueves','viernes','sabado','domingo' }
    Mantiene el orden de aparición en el Excel (fila a fila).
    """
    wb = openpyxl.load_workbook(desde_excel, data_only=True)
    rows: List[Dict[str, Any]] = []

    for sheet in wb.sheetnames:
        if sheet.lower() in ["datos de validación", "datos de validacion", "sustituciones"]:
            continue  # ignorar
        ws = wb[sheet]
        # Cabeceras esperadas: A=Semana, B=Empleado, C..I = L..D
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            semana, empleado = row[0], row[1]
            if empleado is None and semana is None:
                continue
            # Construir dict respetando orden
            d = {
                "hotel": sheet,
                "semana": to_safe_str(semana),
                "empleado": to_safe_str(empleado),
                "lunes": to_safe_str(row[2] if len(row) > 2 else ""),
                "martes": to_safe_str(row[3] if len(row) > 3 else ""),
                "miercoles": to_safe_str(row[4] if len(row) > 4 else ""),
                "jueves": to_safe_str(row[5] if len(row) > 5 else ""),
                "viernes": to_safe_str(row[6] if len(row) > 6 else ""),
                "sabado": to_safe_str(row[7] if len(row) > 7 else ""),
                "domingo": to_safe_str(row[8] if len(row) > 8 else ""),
            }
            # Normalizar NaN/None a cadena vacía
            for k,v in list(d.items()):
                d[k] = to_safe_str(v)
            rows.append(d)
    return rows

def leer_sustituciones_desde_csv(path_csv: str) -> List[Dict[str, Any]]:
    """
    Lee sustituciones ÚNICAMENTE desde CSV. Soporta separador coma o punto y coma.
    Devuelve lista de dicts (no se asume esquema rígido, solo se carga tal cual).
    """
    if not os.path.exists(path_csv):
        print(f"⚠️  CSV de sustituciones no encontrado: {path_csv}")
        return []
    import csv
    # Detectar separador
    with open(path_csv, "r", encoding="utf-8-sig") as f:
        head = f.read(2048)
        sep = ";" if head.count(";") > head.count(",") else ","
    out: List[Dict[str, Any]] = []
    with open(path_csv, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=sep)
        for r in reader:
            out.append({k: (v or "").strip() for k,v in r.items()})
    return out

def reemplazar_data_en_html(html_path: Path, data_obj: Dict[str, Any]) -> bool:
    """
    Abre el HTML y sustituye el bloque:
      const DATA = ...;
    respetando las marcas del template.
    """
    if not html_path.exists():
        print(f"❌ No encuentro {html_path}")
        return False
    src = html_path.read_text(encoding="utf-8")

    import re, json
    # Buscar el patrón const DATA = ... (hasta fin de línea o ;)
    pattern = r"const\s+DATA\s*=\s*[\s\S]*?;"
    new_block = "const DATA = " + json.dumps(data_obj, ensure_ascii=False) + ";"
    if re.search(pattern, src):
        dst = re.sub(pattern, new_block, src, count=1)
    else:
        # Si no existe, lo añadimos antes de </script>
        dst = src.replace("// -------------------------------------------------------------------", "// -------------------------------------------------------------------\n" + new_block)

    html_path.write_text(dst, encoding="utf-8")
    return True

def construir_data(rows: List[Dict[str, Any]], sustituciones_csv: List[Dict[str, Any]]) -> Dict[str, Any]:
    return {
        "rows": rows,
        "meta": {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "rows": len(rows),
            "sustituciones_desde_csv": len(sustituciones_csv),
            "source": "excel+csv",
        },
        "sustituciones": sustituciones_csv  # por si el front lo necesita
    }

def main():
    excel = localizar_excel()
    rows = leer_hoteles_y_rows(excel)
    sustituciones = leer_sustituciones_desde_csv(CSV_SUSTITUCIONES)  # FORZADO CSV
    data = construir_data(rows, sustituciones)

    html = Path("index.html")
    ok = reemplazar_data_en_html(html, data)
    if ok:
        print(f"OK -> index.html actualizado (DATA.rows={len(rows)}, sustituciones_csv={len(sustituciones)})")

if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
Genera/reescribe 'sustituciones_diagnostico.csv' leyendo la hoja 'Sustituciones'
**forzando la lectura completa** con openpyxl (por si Pandas recorta el rango usado).

- No crea carpetas nuevas ni HTML.
- Maneja bloqueo OneDrive (copia temporal si está en uso).
- Parseo robusto de fechas en español.
- Autodetecta la fila de cabecera (si no es la 1) y las columnas con tolerancia.
"""

import os, re, unicodedata, tempfile, time, shutil
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import load_workbook

# =====================
# CONFIGURACIÓN
# =====================
EXCEL_PATH = r"C:\Users\comun\OneDrive\02. Comp. Min Recepción\3. Turnos\Plantilla Cuadrante con Sustituciones v.6.0.xlsx"
SHEET_NAME = "Sustituciones"
OUT_CSV    = r"C:\Users\comun\OneDrive\02. Comp. Min Recepción\3. Turnos\sustituciones_diagnostico.csv"
PAUSE_ON_EXIT = False

TOL_HEADER_ROWS = 10   # Buscamos la cabecera en las 10 primeras filas

# =====================
# UTILIDADES
# =====================
def _canon(s):
    if s is None:
        return ""
    if isinstance(s, float) and np.isnan(s):
        return ""
    return str(s).strip()

_MESES = {"ene":"01","feb":"02","mar":"03","abr":"04","may":"05","jun":"06","jul":"07","ago":"08","sep":"09","set":"09","oct":"10","nov":"11","dic":"12"}
_DOW_RE = re.compile(r'^(lu|lun|ma|mar|mi|mie|mié|ju|jue|vi|vie|sa|sab|sáb|do|dom)[\.\s]+', re.IGNORECASE)

def _strip_accents(s: str) -> str:
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def _parse_fecha(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "", ""
    orig = str(v)
    # datetime real de Excel
    if hasattr(v, 'year'):
        try:
            dt = pd.to_datetime(v, dayfirst=True, errors="coerce")
            return orig, ("" if pd.isna(dt) else dt.strftime("%Y-%m-%d"))
        except Exception:
            pass
    s = _strip_accents(orig).lower().strip()
    s = _DOW_RE.sub("", s)
    s = s.replace(" de ", " ").replace(" del ", " ").replace("-", "/").replace(".", "/")
    s = re.sub(r"\s+", " ", s)
    m = re.search(r'(\d{1,2})[\/\s]([a-z]{3,4})[\/\s](\d{2,4})', s)
    if m:
        dd = int(m.group(1)); mon_txt = m.group(2)[:3]; yy = int(m.group(3))
        if yy < 100: yy += 2000
        mon = _MESES.get(mon_txt, None)
        if mon: return orig, f"{yy:04d}-{mon}-{dd:02d}"
    m2 = re.search(r'(\d{1,2})\/(\d{1,2})\/(\d{2,4})', s)
    if m2:
        dd = int(m2.group(1)); mm = int(m2.group(2)); yy = int(m2.group(3))
        if yy < 100: yy += 2000
        return orig, f"{yy:04d}-{mm:02d}-{dd:02d}"
    dt = pd.to_datetime(orig, dayfirst=True, errors="coerce")
    return orig, ("" if pd.isna(dt) else dt.strftime("%Y-%m-%d"))

def _open_copy_if_locked(path, max_tries=5, wait=0.6):
    last_err = None
    for _ in range(max_tries):
        try:
            with open(path, 'rb'): pass
            return path, None
        except Exception as e:
            last_err = e; time.sleep(wait)
    # copia temporal si sigue bloqueado
    tmp_path = os.path.join(tempfile.gettempdir(), f"_tmp_excel_{int(time.time()*1000)}.xlsx")
    for _ in range(max_tries):
        try:
            shutil.copy2(path, tmp_path)
            return tmp_path, tmp_path
        except Exception as e:
            last_err = e; time.sleep(wait)
    raise last_err if last_err else RuntimeError("No se pudo abrir/copiar el Excel.")

def _normalize(s):
    s = _canon(s).lower()
    s = _strip_accents(s)
    s = s.replace("  "," ").strip()
    return s

def _match_colnames(header_cells):
    # Tolerancia de nombres de columnas
    targets = {
        "hotel": ["hotel"],
        "fecha": ["fecha"],
        "empleado": ["empleado"],
        "cambio": ["cambio de turno","cambio","cambio turno"],
        "sustituto": ["sustituto","sustituye","sustitucion","sustitución"],
        "tipo": ["tipoausencia","tipo ausencia","ausencia","motivo"]
    }
    idx = {}
    for j, c in enumerate(header_cells):
        name = _normalize(c.value)
        for key, alts in targets.items():
            if name in alts and key not in idx:
                idx[key] = j
    ok = all(k in idx for k in ["hotel","fecha","empleado","cambio","sustituto","tipo"])
    return ok, idx

def _find_header(ws):
    # Busca la fila de cabecera en las primeras TOL_HEADER_ROWS filas
    for i in range(1, TOL_HEADER_ROWS+1):
        cells = list(ws[i])
        ok, idx = _match_colnames(cells)
        if ok:
            return i, idx
    # fallback: asumir primera fila
    cells = list(ws[1])
    ok, idx = _match_colnames(cells)
    if ok:
        return 1, idx
    raise RuntimeError("No se localizaron las columnas base en las primeras filas.")

# =====================
# PROCESO PRINCIPAL
# =====================
def main():
    if not os.path.exists(EXCEL_PATH):
        print("❌ No se encuentra el Excel:", EXCEL_PATH); return

    src_path, tmp = _open_copy_if_locked(EXCEL_PATH)
    try:
        wb = load_workbook(src_path, data_only=True, read_only=True)
        if SHEET_NAME not in wb.sheetnames:
            print(f"❌ La hoja '{SHEET_NAME}' no existe en el Excel."); return
        ws = wb[SHEET_NAME]

        header_row, col_idx = _find_header(ws)
        first_data = header_row + 1
        last_row = ws.max_row

        rows_out = []
        vacios_fecha = 0

        for i in range(first_data, last_row+1):
            row = [ws.cell(row=i, column=j+1).value for j in range(len(ws[header_row]))]
            get = lambda key: row[col_idx[key]] if col_idx[key] < len(row) else None
            hotel     = _canon(get("hotel"))
            fecha_txt = get("fecha")
            empleado  = _canon(get("empleado"))
            cambio    = _canon(get("cambio"))
            sustituto = _canon(get("sustituto"))
            tipo      = _canon(get("tipo"))

            # si toda la fila está vacía, saltar
            if not any([hotel, fecha_txt, empleado, cambio, sustituto, tipo]):
                continue

            f_orig, f_norm = _parse_fecha(fecha_txt)
            if not f_norm: vacios_fecha += 1

            # tipo interpretado (simple)
            t = tipo.lower()
            if "vaca" in t: tipo_i = "Vacaciones"
            elif "baja" in t or " it" in t or t == "it": tipo_i = "Baja"
            elif "perm" in t: tipo_i = "Permiso"
            elif "form" in t or "curso" in t: tipo_i = "Formación"
            elif "fest" in t: tipo_i = "Festivo"
            elif "lib" in t: tipo_i = "Libranza"
            elif "desc" in t: tipo_i = "Descanso"
            else: tipo_i = "Cambio de turno" if cambio else ""

            rows_out.append({
                "Hotel": hotel,
                "FechaOriginal": f_orig,
                "FechaNorm": f_norm,
                "Empleado": empleado,
                "CambioDeTurno": cambio,
                "Sustituto": sustituto,
                "TipoAusencia": tipo,
                "TipoInterpretado": tipo_i,
                "ExisteFechaEnHotel": bool(f_norm),
                "ExisteTitular": "",
                "ExisteOtro": "",
                "OtroTexto": "",
                "Motivo": tipo_i,
            })

        out = pd.DataFrame(rows_out, columns=[
            "Hotel","FechaOriginal","FechaNorm","Empleado","CambioDeTurno","Sustituto",
            "TipoAusencia","TipoInterpretado","ExisteFechaEnHotel","ExisteTitular",
            "ExisteOtro","OtroTexto","Motivo"
        ])
        out.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")
        print(f"✅ Reescrito: {OUT_CSV}  (filas exportadas: {len(out)}, últimas filas leídas hasta la fila Excel #{last_row})")
        print(f"   Cabecera detectada en fila: {header_row}. Filas vacías descartadas (sin datos en las 6 columnas): {last_row - header_row - len(out)}")
        if vacios_fecha:
            print(f"   Aviso: {vacios_fecha} filas con fecha sin parsear (se exportan con FechaNorm vacía).")
    finally:
        try: wb.close()
        except Exception: pass
        if tmp and os.path.exists(tmp):
            try: os.remove(tmp)
            except Exception: pass

if __name__ == "__main__":
    start_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    LOG_PATH = os.path.join(os.path.dirname(OUT_CSV), "sustituciones_diagnostico_log.txt")
    try:
        print("⏱ Inicio:", start_ts)
        main()
        print("✔️ Fin correcto")
        with open(LOG_PATH, "a", encoding="utf-8") as lf:
            lf.write("[{}] OK - Ejecutado sin errores.\n".format(start_ts))
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print("❌ Error durante la ejecución:\n" + tb)
        try:
            with open(LOG_PATH, "a", encoding="utf-8") as lf:
                lf.write("[{}] ERROR - {}\n".format(start_ts, tb))
        except Exception:
            pass
    finally:
        if PAUSE_ON_EXIT:
            try: input("\nPulsa ENTER para cerrar…")
            except Exception: pass
